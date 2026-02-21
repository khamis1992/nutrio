from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# Load the workbook
wb = load_workbook("C:/Users/khamis/Documents/nutrio-fuel/docs/plans/Nutrio-Fuel-E2E-Test-Plan.xlsx")

# Define portal sheets
portal_sheets = [
    "Customer Tests",
    "Partner Tests", 
    "Admin Tests",
    "Driver Tests",
    "System Tests"
]

print("Fixing Status column formulas...")

for sheet_name in portal_sheets:
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found, skipping...")
        continue
    
    ws = wb[sheet_name]
    
    # Find the column indices
    # After adding Link column, the order should be:
    # A: TC ID, B: Portal, C: Module, D: Feature, E: Test Case, F: Steps, G: Link
    # H: Expected Result, I: Status, J: Actual Result, K: Notes, L: Priority, M: Tester
    
    status_col = 9  # Column I (9th column)
    expected_col = 8  # Column H (8th column)
    actual_col = 10  # Column J (10th column)
    
    print(f"\nProcessing {sheet_name}:")
    print(f"  Status column: {status_col} (column I)")
    print(f"  Expected column: {expected_col} (column H)")
    print(f"  Actual column: {actual_col} (column J)")
    
    # Update Status column formula for all data rows
    formula_count = 0
    for row in range(2, ws.max_row + 1):
        # Get the status cell
        status_cell = ws.cell(row=row, column=status_col)
        
        # Set the formula: If Actual is empty, show empty. If Actual equals Expected, show PASS, else FAIL
        formula = f'=IF(J{row}="","",IF(J{row}=H{row},"PASS","FAIL"))'
        status_cell.value = formula
        
        # Style for pass/fail will be applied by Excel when calculated
        formula_count += 1
    
    print(f"  Updated {formula_count} rows with formulas")

# Save workbook
wb.save("C:/Users/khamis/Documents/nutrio-fuel/docs/plans/Nutrio-Fuel-E2E-Test-Plan.xlsx")

print("\n" + "="*50)
print("SUCCESS!")
print("="*50)
print("\nStatus column formulas have been fixed in all portal tabs.")
print("\nThe formula is: =IF(J{row}=\"\",\"\",IF(J{row}=H{row},\"PASS\",\"FAIL\"))")
print("\nColumn mapping:")
print("  H: Expected Result")
print("  I: Status (formula)")
print("  J: Actual Result")
print("\nHow it works:")
print("  - If Actual Result is empty → Status shows nothing")
print("  - If Actual Result = Expected Result → Status shows PASS")
print("  - If Actual Result ≠ Expected Result → Status shows FAIL")
print("\nNote: Excel calculates formulas automatically when you open the file.")
print("If formulas don't calculate immediately, press F9 to force calculation.")
