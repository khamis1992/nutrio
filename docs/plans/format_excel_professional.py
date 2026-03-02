#!/usr/bin/env python3
"""
Professional Formatting for Nutrio Fuel E2E Test Plan
Makes the Excel file look polished and professional
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

file_path = r'C:\Users\khamis\Documents\nutrio-fuel-new\docs\plans\Nutrio-Fuel-E2E-Test-Plan.xlsx'

print("=" * 80)
print("PROFESSIONAL FORMATTING - NUTRIO FUEL E2E TEST PLAN")
print("=" * 80)

# Load workbook
wb = load_workbook(file_path)

# Define professional styles
# Colors
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Dark blue
SUBHEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")  # Medium blue
CRITICAL_FILL = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")  # Red
HIGH_FILL = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")  # Orange
MEDIUM_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Yellow
LOW_FILL = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Green
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red

# Fonts
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
DATA_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(bold=True, name="Calibri", size=10)

# Alignment
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Borders
THIN_BORDER = Border(
    left=Side(style='thin', color='B4B4B4'),
    right=Side(style='thin', color='B4B4B4'),
    top=Side(style='thin', color='B4B4B4'),
    bottom=Side(style='thin', color='B4B4B4')
)

HEADER_BORDER = Border(
    left=Side(style='medium', color='1F4E79'),
    right=Side(style='medium', color='1F4E79'),
    top=Side(style='medium', color='1F4E79'),
    bottom=Side(style='medium', color='1F4E79')
)

# Process each sheet
sheet_names = ['Customer Tests', 'Admin Tests', 'Partner Tests', 'Driver Tests', 'System Tests']

for sheet_name in sheet_names:
    if sheet_name not in wb.sheetnames:
        continue
        
    ws = wb[sheet_name]
    print(f"\nFormatting: {sheet_name}")
    
    # Format header row (row 1)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.border = HEADER_BORDER
    
    # Set column widths
    column_widths = {
        'A': 10,   # TC ID
        'B': 12,   # Portal
        'C': 16,   # Module
        'D': 22,   # Feature
        'E': 35,   # Test Case
        'F': 50,   # Steps
        'G': 40,   # Link/URL
        'H': 45,   # Expected Result
        'I': 10,   # Status
        'J': 15,   # Actual Result
        'K': 20,   # Notes
        'L': 12,   # Priority
        'M': 12,   # Tester
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Format data rows
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        for cell in row:
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            
            # Alignment based on column
            if cell.column in [1, 9, 12, 13]:  # TC ID, Status, Priority, Tester
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = LEFT_ALIGN
            
            # Priority color coding (column L = 12)
            if cell.column == 12 and cell.value:
                priority = str(cell.value).strip()
                if priority == "Critical":
                    cell.fill = CRITICAL_FILL
                    cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
                elif priority == "High":
                    cell.fill = HIGH_FILL
                    cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
                elif priority == "Medium":
                    cell.fill = MEDIUM_FILL
                    cell.font = Font(bold=True, name="Calibri", size=10)
                elif priority == "Low":
                    cell.fill = LOW_FILL
                    cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
            
            # Status color coding (column I = 9)
            if cell.column == 9 and cell.value:
                status = str(cell.value).strip().upper()
                if status == "PASS":
                    cell.fill = PASS_FILL
                    cell.font = Font(bold=True, color="006100", name="Calibri", size=10)
                elif status == "FAIL":
                    cell.fill = FAIL_FILL
                    cell.font = Font(bold=True, color="9C0006", name="Calibri", size=10)
    
    # Set row height for header
    ws.row_dimensions[1].height = 30
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Add alternating row colors for better readability
    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for col_idx in range(1, 14):
                cell = ws.cell(row=row_idx, column=col_idx)
                if not cell.fill or cell.fill.start_color.rgb == '00000000':
                    cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    print(f"  > Formatted {ws.max_row - 1} test cases")

# Format Overview sheet if exists
if 'Overview' in wb.sheetnames:
    ws = wb['Overview']
    print(f"\nFormatting: Overview")
    
    # Format title
    ws['A1'].font = Font(bold=True, size=16, color="1F4E79", name="Calibri")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    
    print(f"  > Overview formatted")

# Save workbook
wb.save(file_path)

print("\n" + "=" * 80)
print("PROFESSIONAL FORMATTING COMPLETE!")
print("=" * 80)
print("\nApplied formatting:")
print("  > Professional header styling (dark blue with white text)")
print("  > Optimized column widths for readability")
print("  > Priority color coding (Critical=Red, High=Orange, Medium=Yellow, Low=Green)")
print("  > Status color coding (Pass=Green, Fail=Red)")
print("  > Borders and alignment")
print("  > Alternating row colors")
print("  > Freeze panes for header row")
print("  > Professional fonts (Calibri)")
print(f"\nFile saved: {file_path}")
