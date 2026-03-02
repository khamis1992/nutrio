#!/usr/bin/env python3
"""
E2E Test Plan Updater for Nutrio Fuel
Adds missing test cases to the Excel test plan
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

file_path = r'C:\Users\khamis\Documents\nutrio-fuel-new\docs\plans\Nutrio-Fuel-E2E-Test-Plan.xlsx'

# Load existing workbook
wb = load_workbook(file_path)

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
new_test_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# NEW CUSTOMER TESTS - Add missing test cases
new_customer_tests = [
    # Critical Integration Tests
    ["TC296", "Customer", "Integration", "Meal Completion Atomic", "Meal Completion - Atomic Transaction", 
     "1. Schedule a meal\n2. Mark as completed\n3. Verify nutrition logged\n4. Check idempotency on retry", 
     "https://nutrio.me/dashboard", "Meal completion atomic, no double counting", "", "", "", "Critical", ""],
    
    ["TC297", "Customer", "Integration", "Order Status Flow", "Complete Order Status Transition Flow", 
     "1. Place order\n2. Verify Pending->Confirmed->Preparing->Ready->Out for Delivery->Delivered\n3. Check notifications at each step", 
     "https://nutrio.me/orders/[id]", "Order progresses through all statuses correctly", "", "", "", "Critical", ""],
    
    ["TC298", "Customer", "Integration", "Real-time Updates", "Real-time Order Status Updates", 
     "1. Open order tracking\n2. Have restaurant update status\n3. Verify update appears without refresh", 
     "https://nutrio.me/tracking", "Status updates in real-time via WebSocket", "", "", "", "High", ""],
    
    # Subscription Features
    ["TC299", "Customer", "Subscription", "Subscription Wizard", "Complete Subscription Quiz/Wizard", 
     "1. Click 'Find My Plan'\n2. Answer 3 questions\n3. Get recommendation\n4. Navigate to recommended plan", 
     "https://nutrio.me/subscription/wizard", "Recommended plan matches quiz answers", "", "", "", "Medium", ""],
    
    ["TC300", "Customer", "Subscription", "Quota Warning Banner", "View Quota Warning at 75% Usage", 
     "1. Use 75%+ of weekly meals\n2. View dashboard\n3. Verify warning banner appears\n4. Click to upgrade", 
     "https://nutrio.me/dashboard", "Warning banner displayed with upgrade option", "", "", "", "Medium", ""],
    
    ["TC301", "Customer", "Subscription", "Subscription Gate", "Subscription Gate for Non-Subscribers", 
     "1. Login without subscription\n2. Try to view meal details\n3. Verify gate appears\n4. Click 'View Plans'", 
     "https://nutrio.me/meals/[id]", "Gate displayed, redirects to subscription", "", "", "", "High", ""],
    
    ["TC302", "Customer", "Subscription", "Freeze Subscription", "Freeze Subscription with Return Date", 
     "1. Go to subscription settings\n2. Click 'Freeze'\n3. Select return date\n4. Confirm freeze", 
     "https://nutrio.me/subscription", "Subscription frozen, auto-resume scheduled", "", "", "", "Medium", ""],
    
    ["TC303", "Customer", "Subscription", "Rollover Credits", "View Rollover Meal Credits", 
     "1. Go to subscription\n2. View unused meals from previous week\n3. Verify rollover credits displayed\n4. Check expiration date", 
     "https://nutrio.me/subscription", "Rollover credits displayed with expiration", "", "", "", "Low", ""],
    
    # Gamification Features
    ["TC304", "Customer", "Gamification", "Streak Tracking", "View Meal Streak", 
     "1. Complete meals on consecutive days\n2. View dashboard\n3. Verify streak counter\n4. View streak rewards", 
     "https://nutrio.me/dashboard", "Streak count and rewards displayed", "", "", "", "Low", ""],
    
    ["TC305", "Customer", "Gamification", "Milestones", "View Achieved Milestones", 
     "1. Complete milestone criteria\n2. Go to milestones page\n3. View unlocked milestones\n4. View rewards", 
     "https://nutrio.me/progress/milestones", "Milestones displayed with unlock status", "", "", "", "Low", ""],
    
    # AI Features
    ["TC306", "Customer", "AI", "Smart Meal Recommendations", "View AI Meal Recommendations", 
     "1. Complete onboarding\n2. Go to meals page\n3. View 'Recommended for You' section\n4. Verify personalized suggestions", 
     "https://nutrio.me/meals", "AI recommendations based on profile", "", "", "", "High", ""],
    
    ["TC307", "Customer", "AI", "Weekly Meal Planner", "Use AI Weekly Meal Planner", 
     "1. Go to meal planner\n2. Generate weekly plan\n3. Review AI suggestions\n4. Customize and save", 
     "https://nutrio.me/planner", "AI-generated meal plan matches nutrition goals", "", "", "", "Medium", ""],
    
    ["TC308", "Customer", "AI", "Smart Adjustments", "Receive Smart Nutrition Adjustments", 
     "1. Log weight weekly\n2. Review AI adjustment suggestions\n3. Accept or reject adjustment\n4. See updated targets", 
     "https://nutrio.me/progress", "Adjustments suggested based on progress", "", "", "", "Medium", ""],
    
    # Progress Tracking
    ["TC309", "Customer", "Progress", "Body Measurements", "Log Body Measurements", 
     "1. Go to progress\n2. Click 'Add Measurements'\n3. Enter weight, body fat, etc.\n4. Save and view chart", 
     "https://nutrio.me/progress/body", "Measurements saved, chart updated", "", "", "", "Medium", ""],
    
    ["TC310", "Customer", "Progress", "Weight Tracking", "Log Daily Weight", 
     "1. Go to weight tracking\n2. Enter current weight\n3. View weight trend chart\n4. Check goal progress", 
     "https://nutrio.me/progress/weight", "Weight logged, trend displayed", "", "", "", "High", ""],
    
    ["TC311", "Customer", "Progress", "Water Intake", "Track Daily Water Intake", 
     "1. Go to dashboard\n2. Click water tracker\n3. Add water amount\n4. View daily progress", 
     "https://nutrio.me/dashboard", "Water intake tracked and saved", "", "", "", "Low", ""],
    
    ["TC312", "Customer", "Progress", "Nutrition Dashboard", "View Detailed Nutrition Dashboard", 
     "1. Go to nutrition dashboard\n2. View macros breakdown\n3. Check weekly averages\n4. View insights", 
     "https://nutrio.me/dashboard/nutrition", "Detailed nutrition analytics displayed", "", "", "", "Medium", ""],
    
    ["TC313", "Customer", "Progress", "Health Score", "View Health Score", 
     "1. Complete daily meals\n2. View dashboard\n3. Check health score\n4. View breakdown of score", 
     "https://nutrio.me/dashboard", "Health score calculated and displayed", "", "", "", "Low", ""],
    
    # Order Features
    ["TC314", "Customer", "Orders", "Delivery Tracking", "Track Delivery on Map", 
     "1. Order out for delivery\n2. Open tracking page\n3. View driver location on map\n4. See ETA updates", 
     "https://nutrio.me/tracking/[id]", "Map shows driver location and route", "", "", "", "High", ""],
    
    ["TC315", "Customer", "Orders", "Order Notifications", "Receive Order Notifications", 
     "1. Place order\n2. Wait for status updates\n3. Receive push/email/WhatsApp notifications\n4. Verify content", 
     "https://nutrio.me/orders", "Notifications received for status changes", "", "", "", "High", ""],
    
    ["TC316", "Customer", "Orders", "Reorder Previous", "Quick Reorder from History", 
     "1. Go to order history\n2. Find previous order\n3. Click 'Reorder'\n4. Items added to new order", 
     "https://nutrio.me/orders", "All items from previous order added", "", "", "", "Medium", ""],
    
    ["TC317", "Customer", "Orders", "Meal Review", "Rate and Review Meal", 
     "1. Complete order\n2. Click 'Rate'\n3. Select stars\n4. Add review text\n5. Upload photo\n6. Submit", 
     "https://nutrio.me/orders/[id]", "Review submitted and displayed", "", "", "", "Medium", ""],
    
    # Meal Features
    ["TC318", "Customer", "Meals", "Meal Add-ons", "Add Meal Add-ons", 
     "1. Select meal\n2. View add-on options\n3. Select add-ons\n4. See price update\n5. Add to order", 
     "https://nutrio.me/meals/[id]", "Add-ons added to meal, price updated", "", "", "", "Medium", ""],
    
    ["TC319", "Customer", "Meals", "Featured Restaurants", "View Featured Restaurants", 
     "1. Go to meals page\n2. View 'Featured' section\n3. Click featured restaurant\n4. See promoted meals", 
     "https://nutrio.me/meals", "Featured restaurants displayed prominently", "", "", "", "Low", ""],
    
    ["TC320", "Customer", "Meals", "VIP Discount", "Apply VIP Discount", 
     "1. Have VIP subscription\n2. Add meals to cart\n3. View discount applied\n4. Complete checkout", 
     "https://nutrio.me/checkout", "VIP discount applied to eligible items", "", "", "", "Medium", ""],
    
    # Wallet Features
    ["TC321", "Customer", "Wallet", "Wallet Payment", "Pay with Wallet", 
     "1. Add funds to wallet\n2. Checkout order\n3. Select 'Pay with Wallet'\n4. Verify balance deduction", 
     "https://nutrio.me/checkout", "Payment processed, wallet balance updated", "", "", "", "High", ""],
    
    ["TC322", "Customer", "Wallet", "Auto-recharge", "Set Up Wallet Auto-recharge", 
     "1. Go to wallet settings\n2. Enable auto-recharge\n3. Set threshold and amount\n4. Save settings", 
     "https://nutrio.me/wallet/settings", "Auto-recharge configured", "", "", "", "Low", ""],
    
    # Affiliate Features
    ["TC323", "Customer", "Affiliate", "Affiliate Dashboard", "View Affiliate Dashboard", 
     "1. Apply for affiliate\n2. Get approved\n3. Go to affiliate page\n4. View stats and earnings", 
     "https://nutrio.me/affiliate", "Affiliate dashboard with stats displayed", "", "", "", "Medium", ""],
    
    ["TC324", "Customer", "Affiliate", "Affiliate Payout", "Request Affiliate Payout", 
     "1. Earn affiliate commission\n2. Go to payouts\n3. Request payout\n4. Verify bank transfer", 
     "https://nutrio.me/affiliate/payouts", "Payout request submitted", "", "", "", "Medium", ""],
    
    # Settings
    ["TC325", "Customer", "Settings", "Notification Preferences", "Manage Notification Preferences", 
     "1. Go to settings\n2. Click Notifications\n3. Toggle channels (Push/Email/WhatsApp)\n4. Save preferences", 
     "https://nutrio.me/settings/notifications", "Preferences saved per category", "", "", "", "Medium", ""],
    
    ["TC326", "Customer", "Settings", "Privacy Settings", "Manage Privacy Settings", 
     "1. Go to privacy settings\n2. Toggle data sharing\n3. Request data export\n4. Review privacy policy", 
     "https://nutrio.me/settings/privacy", "Privacy settings applied", "", "", "", "Low", ""],
    
    # Invoice History
    ["TC327", "Customer", "Billing", "Invoice History", "View Invoice History", 
     "1. Go to invoices\n2. View all invoices\n3. Download PDF\n4. Verify details", 
     "https://nutrio.me/invoices", "All invoices listed, PDF downloadable", "", "", "", "Low", ""],
]

# NEW ADMIN TESTS
new_admin_tests = [
    ["TC328", "Admin", "Analytics", "Retention Analytics", "View Customer Retention Analytics", 
     "1. Go to Analytics\n2. Click Retention\n3. View cohort analysis\n4. Check churn rate", 
     "https://nutrio.me/admin/analytics/retention", "Retention metrics and cohorts displayed", "", "", "", "High", ""],
    
    ["TC329", "Admin", "AI", "AI Engine Monitor", "Monitor AI Engine Performance", 
     "1. Go to AI Monitor\n2. View accuracy metrics\n3. Check recommendation quality\n4. Review adjustments", 
     "https://nutrio.me/admin/ai-monitor", "AI performance metrics displayed", "", "", "", "Medium", ""],
    
    ["TC330", "Admin", "Subscription", "Freeze Management", "Manage Subscription Freezes", 
     "1. Go to Freeze Management\n2. View frozen subscriptions\n3. Edit freeze dates\n4. Cancel freeze if needed", 
     "https://nutrio.me/admin/subscriptions/freeze", "Frozen subscriptions managed", "", "", "", "Medium", ""],
    
    ["TC331", "Admin", "Gamification", "Streak Rewards", "Manage Streak Rewards", 
     "1. Go to Streak Rewards\n2. View current rewards\n3. Edit reward thresholds\n4. Add new rewards", 
     "https://nutrio.me/admin/streak-rewards", "Streak rewards configured", "", "", "", "Low", ""],
    
    ["TC332", "Admin", "Gamification", "Milestones", "Manage Milestones", 
     "1. Go to Milestones\n2. View all milestones\n3. Edit criteria\n4. Toggle active status", 
     "https://nutrio.me/admin/milestones", "Milestones configured", "", "", "", "Low", ""],
    
    ["TC333", "Admin", "Featured", "Featured Restaurants", "Manage Featured Restaurants", 
     "1. Go to Featured\n2. Add featured restaurant\n3. Set priority\n4. Schedule feature period", 
     "https://nutrio.me/admin/featured", "Featured restaurants updated", "", "", "", "Low", ""],
    
    ["TC334", "Admin", "Affiliate", "Affiliate Applications", "Review Affiliate Applications", 
     "1. Go to Affiliate Applications\n2. Review pending apps\n3. Approve or reject\n4. Send notification", 
     "https://nutrio.me/admin/affiliates/applications", "Applications reviewed and processed", "", "", "", "Medium", ""],
    
    ["TC335", "Admin", "Affiliate", "Affiliate Payouts", "Process Affiliate Payouts", 
     "1. Go to Affiliate Payouts\n2. View pending\n3. Process payouts\n4. Mark as paid", 
     "https://nutrio.me/admin/affiliates/payouts", "Affiliate payouts processed", "", "", "", "Medium", ""],
    
    ["TC336", "Admin", "IP Management", "IP Restriction", "Manage IP Restrictions", 
     "1. Go to IP Management\n2. View blocked IPs\n3. Add new restriction\n4. Test access", 
     "https://nutrio.me/admin/ip-management", "IP restrictions applied", "", "", "", "High", ""],
    
    ["TC337", "Admin", "Exports", "Data Exports", "Export Platform Data", 
     "1. Go to Exports\n2. Select data type\n3. Set date range\n4. Generate export", 
     "https://nutrio.me/admin/exports", "Export generated and downloaded", "", "", "", "Low", ""],
]

# NEW PARTNER TESTS
new_partner_tests = [
    ["TC338", "Partner", "AI", "AI Insights", "View AI Business Insights", 
     "1. Go to AI Insights\n2. View demand forecast\n3. Check popular times\n4. Review recommendations", 
     "https://nutrio.me/partner/ai-insights", "AI insights displayed for restaurant", "", "", "", "Medium", ""],
    
    ["TC339", "Partner", "Boost", "Boost Features", "Purchase and Use Boost", 
     "1. Go to Boost\n2. Select boost package\n3. Complete payment\n4. Verify boosted placement", 
     "https://nutrio.me/partner/boost", "Restaurant boosted in listings", "", "", "", "Low", ""],
    
    ["TC340", "Partner", "Addons", "Meal Add-ons", "Manage Meal Add-ons", 
     "1. Go to Add-ons\n2. Create add-on category\n3. Add add-on items\n4. Set prices", 
     "https://nutrio.me/partner/addons", "Add-ons created and available", "", "", "", "Medium", ""],
    
    ["TC341", "Partner", "Earnings", "Detailed Earnings Dashboard", "View Detailed Earnings", 
     "1. Go to Earnings Dashboard\n2. View charts\n3. Filter by date\n4. Export data", 
     "https://nutrio.me/partner/earnings", "Detailed earnings analytics displayed", "", "", "", "High", ""],
]

# NEW SYSTEM TESTS
new_system_tests = [
    ["TC342", "System", "Financial", "Credit Transaction Atomicity", "Test Credit Deduction Atomicity", 
     "1. Create test order\n2. Deduct credit\n3. Verify balance\n4. Attempt double deduction", 
     "https://nutrio.me/api/rpc", "Credit deducted atomically, no double-spending", "", "", "", "Critical", ""],
    
    ["TC343", "System", "Financial", "Commission Enforcement", "Verify 10% Commission Rate", 
     "1. Create order\n2. Check earnings record\n3. Verify 10% commission\n4. Check restaurant payout", 
     "https://nutrio.me/api/rpc", "Commission calculated at exactly 10%", "", "", "", "Critical", ""],
    
    ["TC344", "System", "AI", "Nutrition Calculation Accuracy", "Test BMR/TDEE Calculations", 
     "1. Create profile\n2. Check BMR calculation\n3. Verify TDEE with activity\n4. Check macro ratios", 
     "https://nutrio.me/api/rpc", "Calculations accurate within tolerance", "", "", "", "High", ""],
    
    ["TC345", "System", "AI", "Meal Plan Compliance", "Test Macro Compliance >90%", 
     "1. Generate meal plan\n2. Calculate macro totals\n3. Compare to targets\n4. Verify >90% compliance", 
     "https://nutrio.me/api/rpc", "Meal plan achieves >90% compliance", "", "", "", "High", ""],
    
    ["TC346", "System", "Load", "Concurrent Meal Completion", "Test Concurrent Meal Completions", 
     "1. Simulate 10 concurrent users\n2. Complete same meal\n3. Verify only 1 success\n4. Check idempotency", 
     "https://nutrio.me/api/rpc", "Race conditions prevented", "", "", "", "Critical", ""],
    
    ["TC347", "System", "Load", "Payment Double-spending", "Test Payment Double-spending Prevention", 
     "1. Initiate payment\n2. Attempt 50 concurrent processes\n3. Verify only 1 succeeds\n4. Check balance", 
     "https://nutrio.me/api/rpc", "Double-spending prevented", "", "", "", "Critical", ""],
]

print(f"Adding {len(new_customer_tests)} new Customer tests")
print(f"Adding {len(new_admin_tests)} new Admin tests")  
print(f"Adding {len(new_partner_tests)} new Partner tests")
print(f"Adding {len(new_system_tests)} new System tests")
print(f"Total new tests: {len(new_customer_tests) + len(new_admin_tests) + len(new_partner_tests) + len(new_system_tests)}")

# Read existing sheets to get column names
customer_df = pd.read_excel(file_path, sheet_name='Customer Tests')
print(f"\nExisting Customer columns: {list(customer_df.columns)}")
print(f"Existing Customer tests: {len(customer_df)}")

# Create DataFrames for new tests
customer_new_df = pd.DataFrame(new_customer_tests, columns=list(customer_df.columns))
admin_df = pd.read_excel(file_path, sheet_name='Admin Tests')
admin_new_df = pd.DataFrame(new_admin_tests, columns=list(admin_df.columns))
partner_df = pd.read_excel(file_path, sheet_name='Partner Tests')
partner_new_df = pd.DataFrame(new_partner_tests, columns=list(partner_df.columns))
system_df = pd.read_excel(file_path, sheet_name='System Tests')
system_new_df = pd.DataFrame(new_system_tests, columns=list(system_df.columns))

# Combine existing and new tests
customer_combined = pd.concat([customer_df, customer_new_df], ignore_index=True)
admin_combined = pd.concat([admin_df, admin_new_df], ignore_index=True)
partner_combined = pd.concat([partner_df, partner_new_df], ignore_index=True)
system_combined = pd.concat([system_df, system_new_df], ignore_index=True)

print(f"\nUpdated counts:")
print(f"Customer tests: {len(customer_combined)} (added {len(new_customer_tests)})")
print(f"Admin tests: {len(admin_combined)} (added {len(new_admin_tests)})")
print(f"Partner tests: {len(partner_combined)} (added {len(new_partner_tests)})")
print(f"System tests: {len(system_combined)} (added {len(new_system_tests)})")

# Write back to Excel with ExcelWriter
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    customer_combined.to_excel(writer, sheet_name='Customer Tests', index=False)
    admin_combined.to_excel(writer, sheet_name='Admin Tests', index=False)
    partner_combined.to_excel(writer, sheet_name='Partner Tests', index=False)
    system_combined.to_excel(writer, sheet_name='System Tests', index=False)

print("\nExcel file updated successfully!")
print(f"Total tests in plan: {len(customer_combined) + len(admin_combined) + len(partner_combined) + len(system_combined)}")
